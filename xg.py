import io
import json
from collections import OrderedDict
from datetime import datetime
from pathlib import Path

import numpy as np
import pandas as pd
import streamlit as st
from pulp import (
    LpAffineExpression,
    LpMinimize,
    LpProblem,
    LpStatus,
    LpVariable,
    PULP_CBC_CMD,
    lpSum,
)

# Set page configuration (title + favicon + wide layout)
st.set_page_config(
    page_title="Rekenmodule",
    page_icon=str((Path.cwd() / "euroliquids.png").as_posix()),
    layout="wide",
)

# --- Authentication gate ---
try:
    from auth_local import (
        login_gate,
        sidebar_user_pill,
        require_role,
        show_login_toast,
        logout_button,
        list_all_users,
        update_user_role,
        role_counts,
        is_immutable_admin,
    )
    user = login_gate()
    # Remove persistent sidebar user pill and top logout button; handled below near filters
    # show a brief bottom-right toast after login
    try:
        show_login_toast()
    except Exception:
        pass
except Exception as _auth_exc:
    # If auth module is missing or errors, show a clear message including details
    st.error(f"Authenticatie-module ontbreekt of faalt: {_auth_exc}. Controleer 'auth_local.py' en dependencies.")
    st.stop()

BASE_DIR = Path.cwd()
RECEPTEN_PATH = BASE_DIR / "recepten.json"
TANKS_PATH = BASE_DIR / "tanks.json"
UPLOAD_ARCHIVE = BASE_DIR / "dist" / "recepten_uploads"

RECIPE_SHEET_NAME = "Recepten"
COMPOSITION_SHEET_NAME = "Samenstelling"
TANK_SHEET_NAME = "Tanks"
TEMPLATE_VERSION = "2024.08"

RECIPE_TEMPLATE_COLUMNS = [
    "Recept nr.",
    "Recept ID",
    "Naam",
    "Status",
    "Hoeveelheid",
    "Dichtheid",
    "Max Volume liter",
    "Ingangsdatum",
    "Opmerkingen",
]

COMPOSITION_TEMPLATE_COLUMNS = [
    "Recept nr.",
    "Recept ID",
    "Grondstof",
    "Element",
    "Eenheid",
    "Gemeten Concentratie",
    "Spec",
    "Min",
    "Max",
    "Concentratie Grondstof",
    "Ratio Element",
    "Notities",
]

TANK_TEMPLATE_COLUMNS = [
    "Productie tank(s)",
    "Capaciteit (m3)",
    "%",
    "Max. vulgraad (m3)",
]

# "Recept nr." is optional and serves as a disambiguator alongside "Recept ID".
RECIPE_REQUIRED_COLUMNS = {"Recept ID", "Naam", "Hoeveelheid", "Dichtheid"}
# "Recept nr." in composition is optional; if present it is used along with "Recept ID".
COMPOSITION_REQUIRED_COLUMNS = {
    "Recept ID",
    "Grondstof",
    "Element",
    "Eenheid",
    "Gemeten Concentratie",
    "Spec",
    "Min",
    "Max",
    "Concentratie Grondstof",
    "Ratio Element",
}
TANK_REQUIRED_COLUMNS = {"Productie tank(s)", "Capaciteit (m3)", "%", "Max. vulgraad (m3)"}

NUMERIC_RECIPE_COLUMNS = ["Hoeveelheid", "Dichtheid", "Max Volume liter"]

COMPOSITION_NUMERIC_FIELDS = [
    ("Gemeten Concentratie", "Gemeten_Concentratie", True),
    ("Spec", "Spec", False),
    ("Min", "Min", False),
    ("Max", "Max", False),
    ("Concentratie Grondstof", "Concentratie_Grondstof", False),
    ("Ratio Element", "Ratio_Element", False),
]

DEFAULT_RECIPE_META = {"schema_version": 1, "template_version": TEMPLATE_VERSION}


def is_placeholder_element(element: object) -> bool:
    element_str = str(element).strip().lower() if element is not None else ""
    return element_str in {"", "geen", "nan", "none"}


def ensure_archive_directory() -> None:
    """Zorg dat de map voor geüploade recepten aanwezig is."""

    UPLOAD_ARCHIVE.mkdir(parents=True, exist_ok=True)


def format_timestamp(value, default: str = "-") -> str:
    """Zet ISO-timestamp om naar 'YYYY-MM-DD HH:MM'."""

    if not value:
        return default
    try:
        parsed = pd.to_datetime(value, utc=True)
    except Exception:  # noqa: BLE001 - parsing mag niet crashen
        try:
            parsed = pd.to_datetime(value, errors="coerce")
        except Exception:
            return str(value)

    if pd.isna(parsed):
        return default

    if getattr(parsed, "tzinfo", None) is not None:
        parsed = parsed.tz_convert(None)

    return parsed.strftime("%Y-%m-%d %H:%M")


def load_recepten_data():
    """Lees recepten.json en bied ondersteuning voor legacy-structuren."""

    if not RECEPTEN_PATH.exists():
        return {}, {**DEFAULT_RECIPE_META, "last_updated": None}

    try:
        payload = json.loads(RECEPTEN_PATH.read_text(encoding="utf-8"))
    except json.JSONDecodeError as exc:
        st.error(f"Fout bij het inladen van recepten.json: {exc}")
        st.stop()

    if isinstance(payload, dict) and "recipes" in payload:
        recipes = payload.get("recipes", {})
        meta = {**DEFAULT_RECIPE_META, **payload.get("meta", {})}
    elif isinstance(payload, dict):
        recipes = payload
        meta = {**DEFAULT_RECIPE_META, "bron": "legacy"}
    else:
        st.error("Onbekend formaat van recepten.json. Controleer het bestand.")
        st.stop()

    return recipes, meta


def write_recepten_data(recipes: dict, meta: dict) -> None:
    """Bewaar recepten in een gestandaardiseerde structuur."""

    payload = {
        "meta": {**DEFAULT_RECIPE_META, **meta, "last_updated": datetime.utcnow().isoformat()},
        "recipes": recipes,
    }
    RECEPTEN_PATH.write_text(json.dumps(payload, indent=4, ensure_ascii=False), encoding="utf-8")


def write_tanks_data(tanks: dict[str, dict]) -> None:
    """Schrijf de beschikbare tanks naar tanks.json."""

    TANKS_PATH.write_text(json.dumps(tanks, indent=4, ensure_ascii=False), encoding="utf-8")


def coerce_float(value, context: str, errors: list[str], allow_empty: bool = False) -> float | None:
    """Zet waarden om naar float met foutafhandeling."""

    if value is None or (isinstance(value, float) and np.isnan(value)):
        if allow_empty:
            return 0.0
        errors.append(f"Waarde ontbreekt: {context}")
        return None
    try:
        return float(value)
    except (TypeError, ValueError):
        if allow_empty:
            return 0.0
        errors.append(f"Ongeldige numerieke waarde ({value}) voor {context}")
        return None


def coerce_date(value) -> str | None:
    """Converteer datumwaarden naar ISO-notatie."""

    if value is None or (isinstance(value, float) and np.isnan(value)):
        return None
    if isinstance(value, datetime):
        return value.date().isoformat()
    try:
        parsed = pd.to_datetime(value, errors="coerce")
        return parsed.date().isoformat() if pd.notna(parsed) else str(value)
    except Exception:  # noqa: BLE001 - brede catching om inconsistenties op te vangen
        return str(value)


def generate_recipe_template_bytes() -> bytes:
    """Genereer een ingevroren Excel-sjabloon voor receptenbeheer."""

    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    from openpyxl.worksheet.datavalidation import DataValidation
    from openpyxl.utils import get_column_letter

    wb = Workbook()

    ws_recipes = wb.active
    ws_recipes.title = RECIPE_SHEET_NAME
    ws_recipes.append(RECIPE_TEMPLATE_COLUMNS)
    ws_recipes.freeze_panes = "A2"

    header_fill = PatternFill(start_color="FF004B8D", end_color="FF004B8D", fill_type="solid")
    header_font = Font(color="FFFFFFFF", bold=True)
    for col_idx in range(1, len(RECIPE_TEMPLATE_COLUMNS) + 1):
        cell = ws_recipes.cell(row=1, column=col_idx)
        cell.fill = header_fill
        cell.font = header_font
        ws_recipes.column_dimensions[cell.column_letter].width = 22

    status_validation = DataValidation(
        type="list",
        formula1='"Actief,Gepland,Gearchiveerd"',
        allow_blank=True,
        showErrorMessage=True,
        error="Kies een status uit de lijst.",
    )
    ws_recipes.add_data_validation(status_validation)
    # Add validation dynamically on the "Status" column
    status_col_idx = RECIPE_TEMPLATE_COLUMNS.index("Status") + 1
    status_col_letter = get_column_letter(status_col_idx)
    status_validation.add(f"{status_col_letter}2:{status_col_letter}500")

    ws_comp = wb.create_sheet(COMPOSITION_SHEET_NAME)
    ws_comp.append(COMPOSITION_TEMPLATE_COLUMNS)
    ws_comp.freeze_panes = "A2"

    for col_idx in range(1, len(COMPOSITION_TEMPLATE_COLUMNS) + 1):
        cell = ws_comp.cell(row=1, column=col_idx)
        cell.fill = header_fill
        cell.font = header_font
        ws_comp.column_dimensions[cell.column_letter].width = 24

    ws_tanks = wb.create_sheet(TANK_SHEET_NAME)
    ws_tanks.append(TANK_TEMPLATE_COLUMNS)
    ws_tanks.freeze_panes = "A2"

    for col_idx in range(1, len(TANK_TEMPLATE_COLUMNS) + 1):
        cell = ws_tanks.cell(row=1, column=col_idx)
        cell.fill = header_fill
        cell.font = header_font
        ws_tanks.column_dimensions[cell.column_letter].width = 20

    # Voorzie het %-kolom van een formule die automatisch rekent op basis van capaciteit en max. vulgraad
    for row_idx in range(2, 1002):
        percent_cell = ws_tanks.cell(row=row_idx, column=3)
        percent_cell.value = f'=IF(OR(B{row_idx}="",D{row_idx}=""),"",D{row_idx}/B{row_idx}*100)'
        percent_cell.number_format = "0.##"
        ws_tanks.cell(row=row_idx, column=2).number_format = "0.##"
        ws_tanks.cell(row=row_idx, column=4).number_format = "0.##"

    eenheid_validation = DataValidation(
        type="list",
        formula1='"wt%,mg/kg"',
        allow_blank=False,
        showErrorMessage=True,
        error="Gebruik wt% of mg/kg.",
    )
    ws_comp.add_data_validation(eenheid_validation)
    # Add validation dynamically on the "Eenheid" column
    eenheid_col_idx = COMPOSITION_TEMPLATE_COLUMNS.index("Eenheid") + 1
    eenheid_col_letter = get_column_letter(eenheid_col_idx)
    eenheid_validation.add(f"{eenheid_col_letter}2:{eenheid_col_letter}1000")

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()


def generate_recipe_export_bytes(recipes: dict[str, dict], tanks: dict[str, dict]) -> bytes:
    """Vul het sjabloon met de huidige dataset zodat gebruikers eenvoudig kunnen uitbreiden."""

    from openpyxl import load_workbook

    base_workbook = io.BytesIO(generate_recipe_template_bytes())
    wb = load_workbook(base_workbook)

    ws_recipes = wb[RECIPE_SHEET_NAME]
    ws_comp = wb[COMPOSITION_SHEET_NAME]
    ws_tanks = wb[TANK_SHEET_NAME]

    recipe_row = 2
    composition_row = 2
    tank_row = 2

    for naam, payload in sorted(recipes.items()):
        data = payload.get("Data", {})
        max_volume = payload.get("Max_Volume_liter", payload.get("Max_Volume"))
        recept_id = payload.get("Recept_ID") or naam
        recept_nr = payload.get("Recept_Nr") or payload.get("Recept nr.") or ""

        opmerkingen_val = payload.get("Opmerkingen", "")
        if opmerkingen_val is None or (isinstance(opmerkingen_val, str) and opmerkingen_val.strip().lower() == "nan"):
            opmerkingen_val = ""

        recipe_values = {
            "Recept nr.": recept_nr,
            "Recept ID": recept_id,
            "Naam": naam,
            "Status": payload.get("Status", "Actief"),
            "Hoeveelheid": payload.get("Hoeveelheid"),
            "Dichtheid": payload.get("Dichtheid"),
            "Max Volume liter": max_volume,
            "Ingangsdatum": payload.get("Ingangsdatum"),
            "Opmerkingen": opmerkingen_val,
        }

        for col_idx, kolom in enumerate(RECIPE_TEMPLATE_COLUMNS, start=1):
            ws_recipes.cell(row=recipe_row, column=col_idx, value=recipe_values.get(kolom))
        recipe_row += 1

        if not data:
            continue

        row_count = len(data.get("Grondstof", []))
        for i in range(row_count):
            comp_values = {
                "Recept nr.": recept_nr,
                "Recept ID": recept_id,
                "Grondstof": data.get("Grondstof", [None] * row_count)[i],
                "Element": data.get("Element", [None] * row_count)[i],
                "Eenheid": data.get("Eenheid", [None] * row_count)[i],
                "Gemeten Concentratie": data.get("Gemeten_Concentratie", [None] * row_count)[i],
                "Spec": data.get("Spec", [None] * row_count)[i],
                "Min": data.get("Min", [None] * row_count)[i],
                "Max": data.get("Max", [None] * row_count)[i],
                "Concentratie Grondstof": data.get("Concentratie_Grondstof", [None] * row_count)[i],
                "Ratio Element": data.get("Ratio_Element", [None] * row_count)[i],
                "Notities": data.get("Notities", [""] * row_count)[i] if data.get("Notities") else "",
            }

            for col_idx, kolom in enumerate(COMPOSITION_TEMPLATE_COLUMNS, start=1):
                ws_comp.cell(row=composition_row, column=col_idx, value=comp_values.get(kolom))
            composition_row += 1

    for tank_id, info in sorted(tanks.items()):
        label = str(info.get("Name") or tank_id).strip()

        max_volume = info.get("Max_Volume_m3")
        capaciteit = info.get("Capaciteit_m3")
        percentage = info.get("Max_Vulgraad_pct")

        if capaciteit is None:
            # Probeers op basis van max volume en percentage
            if max_volume is not None and percentage:
                try:
                    capaciteit = float(max_volume) / (float(percentage) / 100.0)
                except (ZeroDivisionError, TypeError, ValueError):
                    capaciteit = max_volume
            else:
                capaciteit = max_volume if max_volume is not None else info.get("Capaciteit")

        if percentage is None and capaciteit and max_volume:
            try:
                percentage = round((float(max_volume) / float(capaciteit)) * 100, 2)
            except (ZeroDivisionError, TypeError, ValueError):
                percentage = 100.0

        if percentage is None:
            percentage = 100.0

        try:
            pct_float = round(float(percentage), 2)
        except (TypeError, ValueError):
            pct_float = 100.0

        if max_volume is None and capaciteit is not None:
            try:
                max_volume = round(float(capaciteit) * (pct_float / 100.0), 4)
            except (TypeError, ValueError):
                max_volume = capaciteit

        ws_tanks.cell(row=tank_row, column=1, value=label if label else str(tank_id))
        cap_cell = ws_tanks.cell(row=tank_row, column=2, value=float(capaciteit) if capaciteit is not None else None)
        cap_cell.number_format = "0.00"
        percent_cell = ws_tanks.cell(row=tank_row, column=3)
        percent_cell.value = f'=IF(OR(B{tank_row}="",D{tank_row}=""),"",D{tank_row}/B{tank_row}*100)'
        percent_cell.number_format = "0.00"
        max_cell = ws_tanks.cell(row=tank_row, column=4, value=float(max_volume) if max_volume is not None else None)
        max_cell.number_format = "0.00"
        tank_row += 1

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()


def parse_uploaded_dataset(file_bytes: bytes) -> tuple[dict, dict, dict, dict, list[str]]:
    """Parset een geüpload werkboek naar het interne JSON-formaat, inclusief tanks."""

    errors: list[str] = []
    excel_handle = io.BytesIO(file_bytes)

    try:
        xls = pd.ExcelFile(excel_handle)
    except ValueError as exc:
        errors.append(f"Bestand kan niet worden geopend als Excel: {exc}")
        return {}, {}, {}, {}, errors

    required_sheets = (RECIPE_SHEET_NAME, COMPOSITION_SHEET_NAME, TANK_SHEET_NAME)
    missing_sheets = [sheet for sheet in required_sheets if sheet not in xls.sheet_names]
    if missing_sheets:
        errors.append(f"Ontbrekende tabbladen: {', '.join(missing_sheets)}")
        return {}, {}, {}, {}, errors

    recepten_df = pd.read_excel(xls, RECIPE_SHEET_NAME).dropna(how="all")
    samenstelling_df = pd.read_excel(xls, COMPOSITION_SHEET_NAME).dropna(how="all")
    tanks_df = pd.read_excel(xls, TANK_SHEET_NAME).dropna(how="all")

    missing_recipe_columns = RECIPE_REQUIRED_COLUMNS - set(recepten_df.columns)
    missing_comp_columns = COMPOSITION_REQUIRED_COLUMNS - set(samenstelling_df.columns)
    missing_tank_columns = TANK_REQUIRED_COLUMNS - set(tanks_df.columns)

    if missing_recipe_columns:
        errors.append(
            "Ontbrekende verplichte kolommen in Recepten-tabblad: " + ", ".join(sorted(missing_recipe_columns))
        )
    if missing_comp_columns:
        errors.append(
            "Ontbrekende verplichte kolommen in Samenstelling-tabblad: " + ", ".join(sorted(missing_comp_columns))
        )
    if missing_tank_columns:
        errors.append(
            "Ontbrekende verplichte kolommen in Tanks-tabblad: " + ", ".join(sorted(missing_tank_columns))
        )

    if errors:
        return {}, {}, {}, {}, errors

    recipes_payload: dict[str, dict] = {}
    tanks_payload: dict[str, dict] = {}
    composition_preview: dict[str, pd.DataFrame] = {}
    # Keep track of unique pairs (Recept nr., Recept ID); if nr is empty, use just ID
    gebruikte_ids: set[tuple[str, str]] = set()
    gebruikte_tank_ids: set[str] = set()

    for idx, row in recepten_df.iterrows():
        row_number = idx + 2  # rekening houden met header
        recept_id = str(row.get("Recept ID", "")).strip()
        recept_nr_val = row.get("Recept nr.")
        recept_nr = ""
        if pd.notna(recept_nr_val):
            recept_nr = str(recept_nr_val).strip()
        naam = str(row.get("Naam", "")).strip()

        if not recept_id:
            errors.append(f"Recept ID ontbreekt in Recepten rij {row_number}.")
            continue
        id_pair = (recept_nr, recept_id)
        if id_pair in gebruikte_ids:
            # Exact same pair is a duplicate
            suffix = f" (Recept nr. '{recept_nr}')" if recept_nr else ""
            errors.append(f"Dubbele combinatie Recept ID '{recept_id}'{suffix} in Recepten tabblad.")
            continue
        gebruikte_ids.add(id_pair)

        if not naam:
            errors.append(f"Naam ontbreekt in Recepten rij {row_number}.")
            continue
        if naam in recipes_payload:
            errors.append(f"Dubbele naam '{naam}' in Recepten tabblad.")
            continue

        hoeveelheid = coerce_float(row.get("Hoeveelheid"), f"Hoeveelheid (rij {row_number})", errors)
        dichtheid = coerce_float(row.get("Dichtheid"), f"Dichtheid (rij {row_number})", errors)

        if hoeveelheid is None or dichtheid is None:
            continue

        max_volume_val = row.get("Max Volume liter")
        max_volume = None
        if pd.notna(max_volume_val):
            max_volume = coerce_float(max_volume_val, f"Max Volume liter (rij {row_number})", errors)

        opmerkingen_raw = row.get("Opmerkingen", "")
        if pd.isna(opmerkingen_raw):
            opmerkingen_clean = ""
        else:
            opmerkingen_clean = str(opmerkingen_raw).strip()
            if opmerkingen_clean.lower() == "nan":
                opmerkingen_clean = ""

        recipe_meta = {
            "Recept_ID": recept_id,
            "Recept_Nr": recept_nr or None,
            "Status": str(row.get("Status", "Actief")).strip() or "Actief",
            "Ingangsdatum": coerce_date(row.get("Ingangsdatum")),
            "Opmerkingen": opmerkingen_clean,
        }

        # Match composition rows: prefer (Recept nr., Recept ID) if nr present; otherwise fall back to only ID
        id_matches = samenstelling_df["Recept ID"].astype(str).str.strip() == recept_id
        if "Recept nr." in samenstelling_df.columns and recept_nr:
            nr_series = samenstelling_df["Recept nr."].astype(str).str.strip()
            matching_comp = samenstelling_df[id_matches & (nr_series == recept_nr)]
        else:
            matching_comp = samenstelling_df[id_matches]
        if matching_comp.empty:
            errors.append(f"Geen regels in tabblad Samenstelling voor recept '{naam}'.")
            continue

        comp_dict = {
            "Grondstof": [],
            "Element": [],
            "Eenheid": [],
            "Gemeten_Concentratie": [],
            "Spec": [],
            "Min": [],
            "Max": [],
            "Concentratie_Grondstof": [],
            "Ratio_Element": [],
            "Notities": [],
        }

        for comp_idx, comp_row in matching_comp.iterrows():
            comp_row_num = comp_idx + 2
            grondstof = str(comp_row.get("Grondstof", "")).strip()
            element = str(comp_row.get("Element", "")).strip() or "Geen"
            eenheid = str(comp_row.get("Eenheid", "")).strip()

            if not grondstof:
                errors.append(
                    f"Grondstof ontbreekt in Samenstelling rij {comp_row_num} voor recept '{naam}'."
                )
                continue
            if not eenheid:
                errors.append(
                    f"Eenheid ontbreekt in Samenstelling rij {comp_row_num} voor recept '{naam}'."
                )
                continue

            numeric_values: dict[str, float] = {}
            missing_required = False
            for kolomnaam, intern, optional in COMPOSITION_NUMERIC_FIELDS:
                raw_value = coerce_float(
                    comp_row.get(kolomnaam),
                    f"{kolomnaam} (recept '{naam}', rij {comp_row_num})",
                    errors,
                    allow_empty=optional,
                )
                if raw_value is None:
                    if optional:
                        raw_value = 0.0
                    else:
                        missing_required = True
                        break
                numeric_values[intern] = raw_value

            if missing_required:
                continue

            comp_dict["Grondstof"].append(grondstof)
            comp_dict["Element"].append(element)
            comp_dict["Eenheid"].append(eenheid)
            for intern_key, value in numeric_values.items():
                comp_dict[intern_key].append(value)

            notities_val = comp_row.get("Notities", "")
            if pd.isna(notities_val):
                notities_val = ""
            else:
                notities_val = str(notities_val).strip()
                if notities_val.lower() == "nan":
                    notities_val = ""
            comp_dict["Notities"].append(notities_val)

        # Zorg dat alle lijsten dezelfde lengte hebben
        lengths = {key: len(values) for key, values in comp_dict.items()}
        if not lengths or not any(lengths.values()):
            errors.append(f"Geen geldige samenstellingsregels gevonden voor recept '{naam}'.")
            continue
        if len(set(lengths.values())) != 1:
            errors.append(
                f"Inconsistente kolomlengtes voor recept '{naam}': {lengths}. Controleer op lege cellen."
            )
            continue

        recipes_payload[naam] = {
            "Hoeveelheid": hoeveelheid,
            "Dichtheid": dichtheid,
            "Max_Volume": max_volume,
            "Data": comp_dict,
            **recipe_meta,
        }
        composition_preview[naam] = matching_comp.reset_index(drop=True)

    for idx, row in tanks_df.iterrows():
        row_number = idx + 2
        tank_label = str(row.get("Productie tank(s)", "")).strip()

        if not tank_label:
            errors.append(f"Productie tank ontbreekt in Tanks rij {row_number}.")
            continue
        if tank_label in gebruikte_tank_ids:
            errors.append(f"Dubbele tanknaam '{tank_label}' in Tanks tabblad.")
            continue
        gebruikte_tank_ids.add(tank_label)

        capaciteit = coerce_float(row.get("Capaciteit (m3)"), f"Capaciteit (m3) (rij {row_number})", errors)
        percentage_cell = row.get("%")
        percentage = None
        if pd.notna(percentage_cell) and str(percentage_cell).strip() != "":
            percentage = coerce_float(percentage_cell, f"% (rij {row_number})", errors)
        max_vulgraad_cell = row.get("Max. vulgraad (m3)")
        max_vulgraad = None
        if pd.notna(max_vulgraad_cell) and str(max_vulgraad_cell).strip() != "":
            max_vulgraad = coerce_float(max_vulgraad_cell, f"Max. vulgraad (m3) (rij {row_number})", errors)

        if capaciteit is None:
            continue
        if capaciteit <= 0:
            errors.append(f"Capaciteit (m3) moet groter dan 0 in Tanks rij {row_number}.")
            continue

        if percentage is not None:
            if percentage < 0 or percentage > 100:
                errors.append(f"% (rij {row_number}) moet tussen 0 en 100 liggen.")
                continue

        if percentage is None and max_vulgraad is not None and capaciteit:
            percentage = (max_vulgraad / capaciteit) * 100

        if percentage is None:
            percentage = 100.0

        expected_max = capaciteit * (percentage / 100.0)

        if max_vulgraad is None:
            max_vulgraad = expected_max
        elif max_vulgraad <= 0:
            errors.append(f"Max. vulgraad (m3) moet groter dan 0 in Tanks rij {row_number}.")
            continue
        else:
            if abs(expected_max - max_vulgraad) > max(0.01, 0.005 * capaciteit):
                errors.append(
                    f"Max. vulgraad (m3) ({max_vulgraad}) komt niet overeen met capaciteit * % ({expected_max:.2f}) in Tanks rij {row_number}."
                )
                continue

        percentage = round(float(percentage), 2)
        max_vulgraad = round(float(max_vulgraad), 4)

        tanks_payload[tank_label] = {
            "Name": tank_label,
            "Capaciteit_m3": capaciteit,
            "Max_Vulgraad_pct": percentage,
            "Max_Volume_m3": max_vulgraad,
        }

    return (
        recipes_payload,
        tanks_payload,
        {"recepten": recepten_df, "samenstelling": samenstelling_df, "tanks": tanks_df},
        composition_preview,
        errors,
    )


def render_recipe_management(recipes: dict[str, dict], tanks: dict[str, dict], meta: dict) -> None:
    """UI voor het beheren van recepten via Excel-sjablonen."""

    st.title("Receptbeheer")
    st.markdown(
        """
        Download het sjabloon, vul of wijzig de gegevens, en upload het bestand om de recepten voor
        de rekenmodule centraal te actualiseren. Alle teamleden zien direct dezelfde dataset.
        """
    )

    st.subheader("Dataset exporteren")
    export_bytes = generate_recipe_export_bytes(recipes, tanks)
    st.download_button(
        label="Download huidige dataset (Excel)",
        data=export_bytes,
        file_name="correctierecepten_actueel.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        help="Gebruik dit bestand als uitgangspunt om recepten of tanks te wijzigen of uit te breiden.",
    )
    if not recipes:
        st.info("Nog geen recepten beschikbaar om te exporteren.")

    st.markdown(
        """
        **Nieuwe recepten toevoegen**

        1. Voeg op het tabblad *Recepten* een nieuwe rij toe met `Recept nr.` (optioneel), een uniek `Recept ID`, de `Naam` en de basisgegevens.
        2. Vul op het tabblad *Samenstelling* meerdere rijen in met dezelfde `Recept ID` en, indien gebruikt, dezelfde `Recept nr.`; specificeer per grondstof de specs.
        3. Sla het bestand op en upload het via onderstaande sectie.
        """
    )

    st.subheader("Huidige recepten")
    if recipes:
        overzicht_data = []
        laatst_bijgewerkt = format_timestamp(
            meta.get("last_uploaded_at") or meta.get("last_updated")
        )
        for naam, payload in recipes.items():
            recept_id = payload.get("Recept_ID") or naam
            recept_nr = payload.get("Recept_Nr") or ""
            overzicht_data.append(
                {
                    "Recept nr.": recept_nr,
                    "Recept ID": recept_id,
                    "Naam": naam,
                    "Status": payload.get("Status", "Actief"),
                    "Hoeveelheid (L)": payload.get("Hoeveelheid", 0),
                    "Dichtheid (kg/L)": payload.get("Dichtheid", 0),
                    "Max Volume (L)": payload.get("Max_Volume") or payload.get("Max_Volume_liter"),
                    "Laatste Wijziging": laatst_bijgewerkt,
                }
            )
        overzicht_df = pd.DataFrame(overzicht_data)
        st.dataframe(overzicht_df.sort_values("Naam"), use_container_width=True)
    else:
        st.info("Er zijn nog geen recepten beschikbaar in het systeem.")

    st.subheader("Huidige tanks")
    if tanks:
        tank_overzicht = []
        for tank_id, info in sorted(tanks.items()):
            naam = str(info.get("Name") or tank_id)
            max_vulgraad = info.get("Max_Volume_m3")
            capaciteit = info.get("Capaciteit_m3")
            percentage = info.get("Max_Vulgraad_pct")

            if capaciteit is None and max_vulgraad is not None and percentage:
                try:
                    capaciteit = float(max_vulgraad) / (float(percentage) / 100.0)
                except (ZeroDivisionError, TypeError, ValueError):
                    capaciteit = max_vulgraad

            if percentage is None and capaciteit and max_vulgraad:
                try:
                    percentage = round((float(max_vulgraad) / float(capaciteit)) * 100, 2)
                except (ZeroDivisionError, TypeError, ValueError):
                    percentage = 100.0

            if percentage is None:
                percentage = 100.0 if capaciteit else None

            if max_vulgraad is None and capaciteit is not None and percentage is not None:
                try:
                    max_vulgraad = round(float(capaciteit) * (float(percentage) / 100.0), 4)
                except (TypeError, ValueError):
                    max_vulgraad = capaciteit

            tank_overzicht.append(
                {
                    "Productie tank(s)": naam,
                    "Capaciteit (m³)": capaciteit,
                    "%": percentage,
                    "Max. vulgraad (m³)": max_vulgraad,
                }
            )
        st.dataframe(pd.DataFrame(tank_overzicht), use_container_width=True)
    else:
        st.info("Er zijn nog geen tanks beschikbaar in het systeem.")

    st.divider()
    st.subheader("Upload bijgewerkte recepten")
    st.markdown(
        """
        Upload een volledig ingevuld werkboek. Alle recepten in het bestand vervangen de huidige set. Er zullen
        meerdere validatiecontroles plaatsvinden en een bevestiging is vereist voordat het bestand wordt opgeslagen.
        """
    )

    uploaded_file = st.file_uploader(
        "Kies een Excel-bestand (.xlsx)",
        type=["xlsx"],
        help="Alleen sjablonen met de tabbladen Recepten, Samenstelling en Tanks worden geaccepteerd.",
        key="recepten_upload",
    )

    if uploaded_file is not None:
        file_bytes = uploaded_file.getvalue()
        (
            parsed_recipes,
            parsed_tanks,
            preview_frames,
            composition_frames,
            parse_errors,
        ) = parse_uploaded_dataset(file_bytes)

        if parse_errors:
            for err in parse_errors:
                st.error(err)
        else:
            st.success("Upload succesvol gevalideerd. Controleer de gegevens hieronder voordat je opslaat.")

        if preview_frames:
            recepten_preview = preview_frames.get("recepten")
            if recepten_preview is not None:
                st.markdown("**Recepten-tabblad**")
                st.dataframe(recepten_preview, use_container_width=True)

            tanks_preview = preview_frames.get("tanks")
            if tanks_preview is not None:
                st.markdown("**Tanks-tabblad**")
                st.dataframe(tanks_preview, use_container_width=True)

            st.markdown("**Samenstelling per recept**")
            for naam, comp_df in composition_frames.items():
                with st.expander(f"{naam} ({len(comp_df)} regels)", expanded=False):
                    st.dataframe(comp_df, use_container_width=True)

        if not parse_errors:
            huidige_namen = set(recipes.keys())
            nieuwe_namen = set(parsed_recipes.keys())
            toegevoegde = sorted(nieuwe_namen - huidige_namen)
            verwijderd = sorted(huidige_namen - nieuwe_namen)
            potentieel_bijgewerkt = sorted(nieuwe_namen & huidige_namen)
            bijgewerkt = [naam for naam in potentieel_bijgewerkt if parsed_recipes[naam] != recipes.get(naam)]

            huidige_tanks = set(tanks.keys())
            nieuwe_tanks = set(parsed_tanks.keys())
            tank_toegevoegd = sorted(nieuwe_tanks - huidige_tanks)
            tank_verwijderd = sorted(huidige_tanks - nieuwe_tanks)
            tank_bijgewerkt = [
                tank_id for tank_id in sorted(nieuwe_tanks & huidige_tanks)
                if parsed_tanks[tank_id] != tanks.get(tank_id)
            ]

            st.markdown("**Wijzigingsoverzicht**")
            if toegevoegde:
                st.success("Recepten toegevoegd: " + ", ".join(toegevoegde))
            if bijgewerkt:
                st.info("Recepten bijgewerkt: " + ", ".join(bijgewerkt))
            if verwijderd:
                st.warning("Recepten verwijderd: " + ", ".join(verwijderd))
            if tank_toegevoegd:
                st.success("Tanks toegevoegd: " + ", ".join(tank_toegevoegd))
            if tank_bijgewerkt:
                st.info("Tanks bijgewerkt: " + ", ".join(tank_bijgewerkt))
            if tank_verwijderd:
                st.warning("Tanks verwijderd: " + ", ".join(tank_verwijderd))
            if not any([toegevoegde, bijgewerkt, verwijderd, tank_toegevoegd, tank_bijgewerkt, tank_verwijderd]):
                st.write("Geen verschillen ten opzichte van de huidige dataset gevonden.")

            bevestiging = st.checkbox(
                "Ik bevestig dat de huidige recepten en tanks mogen worden vervangen door deze upload.",
                key="confirm_dataset_replace",
            )

            if bevestiging and st.button("Opslaan en delen", key="commit_recipe_upload"):
                timestamp = datetime.utcnow().strftime("%Y%m%d-%H%M%S")
                archive_name = f"recepten_upload_{timestamp}.xlsx"
                archive_path = UPLOAD_ARCHIVE / archive_name
                archive_path.write_bytes(file_bytes)

                total_elements = sum(len(payload["Data"].get("Grondstof", [])) for payload in parsed_recipes.values())
                meta_update = {
                    "template_version": TEMPLATE_VERSION,
                    "last_uploaded_at": datetime.utcnow().isoformat(),
                    "bron": "excel-upload",
                    "archive_file": archive_name,
                    "total_recipes": len(parsed_recipes),
                    "total_componenten": total_elements,
                    "total_tanks": len(parsed_tanks),
                }
                write_recepten_data(parsed_recipes, meta_update)
                write_tanks_data(parsed_tanks)
                st.success("Dataset opgeslagen en gesynchroniseerd. Alle gebruikers zien de nieuwe set.")
                rerun_fn = getattr(st, "experimental_rerun", None)
                if rerun_fn is None:
                    rerun_fn = getattr(st, "rerun", None)
                if callable(rerun_fn):
                    rerun_fn()
                else:
                    st.info("Herlaad de pagina handmatig om de vernieuwingen te zien.")
                    st.stop()

    st.divider()
    st.subheader("Uploadhistoriek")
    archive_files = sorted(UPLOAD_ARCHIVE.glob("*.xlsx"), key=lambda p: p.stat().st_mtime, reverse=True)
    if archive_files:
        historie_data = []
        for path in archive_files[:10]:
            dt_str = datetime.fromtimestamp(path.stat().st_mtime).strftime("%d-%m-%Y %H:%M")
            size_kb = path.stat().st_size / 1024
            size_str = f"{size_kb:.1f}".rstrip("0").rstrip(".")
            historie_data.append({"Bestand": path.name, "Datum": dt_str, "Grootte (KB)": size_str})
        st.table(pd.DataFrame(historie_data))
    else:
        st.info("Er zijn nog geen geüploade versies gearchiveerd.")


def render_user_management(current_user) -> None:
    """Visuele beheerconsole voor het toekennen van rollen aan gebruikers."""

    st.title("Gebruikersbeheer")
    st.markdown(
        """
        Beheer toegangsrechten voor het team. Promoveer gebruikers tot beheerder of zet ze terug naar
        reguliere gebruikers. Wijzigingen worden direct opgeslagen en zijn meteen actief.
        """
    )

    counts = role_counts()
    col_admins, col_users = st.columns(2)
    col_admins.metric("Beheerders", counts.get("admin", 0))
    col_users.metric("Gebruikers", counts.get("user", 0))

    feedback = st.session_state.pop("_user_admin_feedback", None)
    if feedback:
        level = feedback.get("level")
        message = feedback.get("msg", "")
        if level == "success":
            st.success(message)
        elif level == "info":
            st.info(message)
        else:
            st.error(message)

    st.markdown(
        """
        <style>
        .user-admin-wrapper {display: flex; flex-direction: column; gap: 1rem;}
        .user-admin-card {background: #ffffff; border: 1px solid #E5E7EB; border-radius: 12px; padding: 1rem; box-shadow: 0 8px 16px rgba(15, 23, 42, 0.04);}
        .user-admin-card.admin {border-left: 5px solid #2563EB;}
        .user-admin-card.user {border-left: 5px solid #10B981;}
        .user-admin-card h3 {margin: 0; font-size: 1rem; font-weight: 600; color: #111827;}
        .user-admin-card p {margin: 0.35rem 0 0; color: #4B5563;}
        .user-admin-tags {margin-top: 0.75rem; display: flex; gap: 0.35rem; flex-wrap: wrap;}
        .user-admin-tag {display: inline-flex; align-items: center; padding: 0.2rem 0.55rem; border-radius: 999px; font-size: 0.75rem; font-weight: 500;}
        .user-admin-tag.primary {background: rgba(37, 99, 235, 0.12); color: #1D4ED8;}
        .user-admin-tag.success {background: rgba(16, 185, 129, 0.12); color: #059669;}
        .user-admin-tag.warning {background: rgba(217, 119, 6, 0.12); color: #B45309;}
        .user-admin-tag.muted {background: rgba(107, 114, 128, 0.1); color: #374151;}
        div[data-testid="column"] div.stButton > button {width: 100%; border-radius: 10px; padding: 0.6rem 1rem; font-weight: 600;}
        </style>
        """,
        unsafe_allow_html=True,
    )

    users = list_all_users()
    if not users:
        st.info("Nog geen gebruikers geregistreerd.")
        return

    current_email = (current_user.email or "").lower()

    def _apply_role_change(email: str, target_role: str) -> None:
        ok, msg = update_user_role(email, target_role)
        st.session_state["_user_admin_feedback"] = {
            "level": "success" if ok else "error",
            "msg": msg,
        }
        rerun_fn = getattr(st, "experimental_rerun", None) or getattr(st, "rerun", None)
        if callable(rerun_fn):
            rerun_fn()

    st.markdown('<div class="user-admin-wrapper">', unsafe_allow_html=True)
    for idx, info in enumerate(users):
        email = info.get("email", "")
        role = info.get("role", "user")
        verified = bool(info.get("email_verified"))
        pending_verification = bool(info.get("verification_pending"))
        reset_requested = bool(info.get("reset_requested"))

        tags: list[str] = []
        role_label = "Beheerder" if role == "admin" else "Gebruiker"
        tags.append(f'<span class="user-admin-tag primary">{role_label}</span>')
        if verified:
            tags.append('<span class="user-admin-tag success">E-mail geverifieerd</span>')
        else:
            tags.append('<span class="user-admin-tag warning">Verificatie vereist</span>')
        # Intentionally omit extra pills like 'Verificatielink actief' and 'Reset aangevraagd' for a cleaner UI.

        card_html = f"""
        <div class="user-admin-card {'admin' if role == 'admin' else 'user'}">
            <h3>{email}</h3>
            <p>{'Volledige toegang tot beheerfuncties' if role == 'admin' else 'Standaardtoegang tot de rekenmodule'}</p>
            <div class="user-admin-tags">{''.join(tags)}</div>
        </div>
        """

        col_info, col_actions = st.columns([4, 1])
        with col_info:
            st.markdown(card_html, unsafe_allow_html=True)
        with col_actions:
            button_key = f"role_action_{idx}"
            if role == "admin":
                disable_demote = (
                    email.lower() == current_email
                    or counts.get("admin", 0) <= 1
                    or is_immutable_admin(email)
                )
                if st.button(
                    "Maak gebruiker",
                    key=f"demote_{button_key}",
                    use_container_width=True,
                    disabled=disable_demote,
                ):
                    _apply_role_change(email, "user")
                if disable_demote:
                    if is_immutable_admin(email):
                        st.caption("Vaste beheerder kan niet gedegradeerd worden.")
                    elif email.lower() == current_email:
                        st.caption("Log in als een andere beheerder om jezelf te degraderen.")
                    else:
                        st.caption("Minstens één beheerder moet actief blijven.")
            else:
                if st.button(
                    "Maak beheerder",
                    key=f"promote_{button_key}",
                    use_container_width=True,
                ):
                    _apply_role_change(email, "admin")

        st.markdown("<div style=\"height:0.5rem\"></div>", unsafe_allow_html=True)
    st.markdown("</div>", unsafe_allow_html=True)

ensure_archive_directory()
recepten, recepten_meta = load_recepten_data()

# Load tanks from JSON (no fallback)
try:
    tanks = json.loads(TANKS_PATH.read_text(encoding="utf-8"))
except FileNotFoundError:
    st.error("tanks.json ontbreekt. Plaats het bestand in dezelfde map als de app.")
    st.stop()
except json.JSONDecodeError as exc:
    st.error(f"tanks.json bevat ongeldige JSON: {exc}")
    st.stop()

menu_opties = ["Rekenmodule", "Receptbeheer"]
if (user.role or "").lower() == "admin":
    menu_opties.append("Gebruikersbeheer")

modus = st.sidebar.radio("Kies scherm", menu_opties, index=0)

if modus == "Gebruikersbeheer":
    require_role(user, allowed=("admin",))
    render_user_management(user)
    st.stop()

if modus == "Receptbeheer":
    # Restrict recipe management to admins
    require_role(user, allowed=("admin",))
    render_recipe_management(recepten, tanks, recepten_meta)
    st.stop()

# Titel van de app
st.title("Correctie Recept Rekenmodule")

if not recepten:
    st.warning("Er zijn momenteel geen recepten beschikbaar. Voeg recepten toe via Receptbeheer.")
    st.stop()

# Always offer the full status set; default shows only 'Actief'
alle_statusopties = ["Actief", "Gepland", "Gearchiveerd"]
beschikbare_statussen = alle_statusopties
standaard_status = ["Actief"]

geselecteerde_statussen = st.sidebar.multiselect(
    "Toon recepten met status",
    options=beschikbare_statussen,
    default=standaard_status,
    help="Recepten buiten de geselecteerde status(sen) worden verborgen in de rekenmodule."
)

# Place the logout button directly under the filter in the sidebar
with st.sidebar:
    st.markdown("---")
    logout_button("Uitloggen", key="logout_btn_under_filter", _button_fn=st.sidebar.button)

if not geselecteerde_statussen:
    st.warning("Selecteer minimaal één status om recepten te tonen.")
    st.stop()

gefilterde_recepten = {
    naam: data
    for naam, data in recepten.items()
    if ((data.get("Status") or "Actief").strip() or "Actief") in geselecteerde_statussen
}

if not gefilterde_recepten:
    st.warning("Geen recepten beschikbaar voor de gekozen status(sen). Pas de filter aan.")
    st.stop()

recepten = gefilterde_recepten

if not tanks:
    st.error("Er zijn geen tanks beschikbaar. Controleer tanks.json.")
    st.stop()

# Recept- en tankselectie
col_recept, col_tank = st.columns(2)
with col_recept:
    def _recept_label(nm: str) -> str:
        payload = recepten.get(nm, {})
        nr = payload.get("Recept_Nr") or payload.get("Recept nr.") or payload.get("Recept nr") or ""
        return f"{nm} – Recept nr. {nr}" if nr else nm

    recept = st.selectbox(
        "Kies een recept",
        list(recepten.keys()),
        format_func=_recept_label,
        help="Selecteer het recept om te bewerken.",
    )
with col_tank:
    tank = st.selectbox("Kies een tank", list(tanks.keys()), help="Selecteer de tank/mixer voor de berekening.")

# Algemene gegevens
with st.expander("Algemene Gegevens", expanded=True):
    col1, col2 = st.columns(2)
    with col1:
        hoeveelheid = st.number_input("Hoeveelheid (liter)", value=recepten[recept]["Hoeveelheid"], step=100.0, min_value=0.0, help="De totale hoeveelheid product in liters.")
    with col2:
        dichtheid = st.number_input("Dichtheid (kg/l)", value=recepten[recept]["Dichtheid"], step=0.01, min_value=0.01, help="De dichtheid van het product in kg per liter.", format="%.3f")
    massa_product = hoeveelheid * dichtheid
    st.info(f"Originele massa product: **{massa_product:.2f} kg**")

# Toggle voor Max_Volume constraint
col_max_vol1, col_max_vol2 = st.columns([1, 2])
with col_max_vol1:
    use_max_volume = st.checkbox("Gebruik Max. Volume beperking", value=True, help="Schakel uit om de maximale volumebeperking te negeren.")
with col_max_vol2:
    default_max_volume_m3 = tanks[tank]["Max_Volume_m3"]
    max_volume_m3 = st.number_input(
        "Max Volume (m³)",
        min_value=0.0,
        value=default_max_volume_m3,
        step=0.001,
        disabled=not use_max_volume,
        help=f"Maximale volumebeperking voor {tanks[tank]['Name']} in kubieke meters."
    )
    max_volume_liters = max_volume_m3 * 1000.0

# Gegevens uit het geselecteerde recept
data = recepten[recept]["Data"]
df = pd.DataFrame(data)

element_info_map: OrderedDict[str, dict[str, list]] = OrderedDict()
for idx, row in df.iterrows():
    element_key = row.get('Element')
    info = element_info_map.setdefault(
        element_key,
        {"indices": [], "eenheden": [], "grondstoffen": []},
    )
    info["indices"].append(idx)
    info["grondstoffen"].append(row.get('Grondstof'))
    eenheid_val = row.get('Eenheid')
    if pd.notna(eenheid_val):
        eenheid_str = str(eenheid_val)
        if eenheid_str not in info["eenheden"]:
            info["eenheden"].append(eenheid_str)

grondstoffen = list(dict.fromkeys(df['Grondstof']))
grondstof_to_idx = {naam: idx for idx, naam in enumerate(grondstoffen)}
row_to_grondstof_idx = df['Grondstof'].map(grondstof_to_idx).to_numpy(dtype=int)

st.session_state.current_grondstoffen = grondstoffen
st.session_state.row_to_grondstof_idx = row_to_grondstof_idx

# Initialize step sizes and costs in session state
if (
    'stapgroottes' not in st.session_state
    or 'kosten' not in st.session_state
    or st.session_state.get('current_recept') != recept
    or st.session_state.get('current_tank') != tank
):
    st.session_state.stapgroottes = {grondstof: 0 for grondstof in grondstoffen}
    st.session_state.kosten = {
        grondstof: 0.01 if grondstof == 'Demiwater' else 1.0
        for grondstof in grondstoffen
    }
    st.session_state.excluded_grondstoffen = []
    st.session_state.current_recept = recept
    st.session_state.current_tank = tank

# Input for measured concentrations
with st.expander("Gemeten Concentratie per Element", expanded=True):
    st.info("Voer de gemeten concentraties in (wt% of mg/kg, afhankelijk van de eenheid).")
    measured_concentrations = {}
    for element, info in element_info_map.items():
        indices = info["indices"]
        if not indices:
            continue
        first_idx = indices[0]
        eenheden = [val for val in info["eenheden"] if str(val).strip().lower() != "nan"]
        eenheid = eenheden[0] if eenheden else str(df.at[first_idx, 'Eenheid'])
        grondstof_label = ", ".join(dict.fromkeys(info["grondstoffen"]) ) if info["grondstoffen"] else ""

        if is_placeholder_element(element):
            measured_concentrations[element] = 0.0
            st.number_input(
                f"{element} ({eenheid})" + (f" voor {grondstof_label}" if grondstof_label else ""),
                value=0.0,
                disabled=True,
                key=f"conc_{recept}_{tank}_{element}_{first_idx}",
                help="Geen meting vereist voor deze component."
            )
            continue

        if len(eenheden) > 1:
            st.error(
                f"Element {element} heeft meerdere eenheden ({', '.join(eenheden)}). Harmoniseer de invoer."
            )
            st.stop()

        default_conc = float(df.at[first_idx, 'Gemeten_Concentratie'])
        is_wt_pct = eenheid == 'wt%'
        step_value = 0.01 if is_wt_pct else 1.0
        format_str = "%.2f" if is_wt_pct else "%.0f"
        help_text = f"Bronnen: {grondstof_label}" if grondstof_label else None

        conc = st.number_input(
            f"{element} ({eenheid})" + (f" voor {grondstof_label}" if grondstof_label else ""),
            value=default_conc,
            step=step_value,
            min_value=0.0,
            format=format_str,
            key=f"conc_{recept}_{tank}_{element}",
            help=help_text,
        )
        measured_concentrations[element] = conc

    for placeholder_element in element_info_map:
        if is_placeholder_element(placeholder_element):
            measured_concentrations.setdefault(placeholder_element, 0.0)

df['Gemeten Concentratie'] = df['Element'].map(measured_concentrations).fillna(0.0)

element_groups = []
element_actuele_mass_map: dict[str, float] = {}

for element, info in element_info_map.items():
    indices = info["indices"]
    if not indices:
        continue
    if is_placeholder_element(element):
        element_actuele_mass_map[element] = 0.0
        continue

    eenheden = [val for val in info["eenheden"] if str(val).strip().lower() != "nan"]
    eenheid = eenheden[0] if eenheden else str(df.at[indices[0], 'Eenheid'])
    if len(eenheden) > 1:
        st.error(
            f"Element {element} heeft meerdere eenheden ({', '.join(eenheden)}). Harmoniseer de invoer."
        )
        st.stop()

    measured_value = float(measured_concentrations.get(element, 0.0))
    conversie_factor = 100 if eenheid == 'wt%' else 1_000_000
    actuele_mass = (massa_product * measured_value) / conversie_factor if conversie_factor else 0.0
    element_actuele_mass_map[element] = actuele_mass

    spec_series = df.loc[indices, 'Spec'].astype(float)
    min_series = df.loc[indices, 'Min'].astype(float)
    max_series = df.loc[indices, 'Max'].astype(float)

    spec_values = spec_series.dropna().unique()
    min_values = min_series.dropna().unique()
    max_values = max_series.dropna().unique()

    if len(spec_values) > 1:
        st.error(f"Element {element} heeft verschillende Spec-waarden ({spec_values}). Harmoniseer de data.")
        st.stop()
    if len(min_values) > 1:
        st.error(f"Element {element} heeft verschillende Min-waarden ({min_values}). Harmoniseer de data.")
        st.stop()
    if len(max_values) > 1:
        st.error(f"Element {element} heeft verschillende Max-waarden ({max_values}). Harmoniseer de data.")
        st.stop()

    spec_value = float(spec_series.iloc[0]) if not spec_series.empty and pd.notna(spec_series.iloc[0]) else 0.0
    min_value = float(min_series.iloc[0]) if not min_series.empty and pd.notna(min_series.iloc[0]) else 0.0
    max_value = float(max_series.iloc[0]) if not max_series.empty and pd.notna(max_series.iloc[0]) else 0.0

    ratios = df.loc[indices, 'Ratio_Element'].astype(float).to_numpy()
    grondstof_indices = row_to_grondstof_idx[indices].astype(int)

    element_groups.append(
        {
            'element': element,
            'eenheid': eenheid,
            'indices': indices,
            'ratios': ratios,
            'grondstof_indices': grondstof_indices,
            'actuele_mass': actuele_mass,
            'spec': spec_value,
            'min': min_value,
            'max': max_value,
        }
    )

element_actuele_mass_map.setdefault('Geen', 0.0)
element_actuele_mass_map.setdefault('nan', 0.0)


# Optimalisatie Penalties
with st.expander("Optimalisatie Penalties", expanded=True):
    st.info("Stel de verhouding in tussen de penalties voor massaminimalisatie en specificatie afwijkingen. De slider bepaalt de verhouding. Standaard is 1:10")

    # Initialize session state if not already set
    if 'ratio_log' not in st.session_state:
        st.session_state.ratio_log = 1.0  # Default to 1:1 ratio
    if 'mass_penalty' not in st.session_state:
        st.session_state.mass_penalty = 10000.0
    if 'deviation_weight' not in st.session_state:
        st.session_state.deviation_weight = 10000.0

    # Reset button for penalties
    if st.button("Reset Penalties naar Standaard", help="Herstel de standaard penalties (verhouding 1:1)"):
        st.session_state.ratio_log = 1.0
        st.session_state.mass_penalty = 10000.0
        st.session_state.deviation_weight = 10000.0

    # Ratio slider with explicit callback to update session state
    def update_ratio():
        st.session_state.ratio_log = st.session_state.ratio_slider

    ratio_log = st.slider(
        "Straf Verhouding (Massaminimalisatie : Specificatie Afwijkingen)",
        min_value=0.0,  # Represents 1:1 (log10(1))
        max_value=2.0,  # Represents 100:1 (log10(100))
        value=st.session_state.ratio_log,  # Use session state for slider value
        step=0.1,
        format="%.1f",
        key="ratio_slider",
        on_change=update_ratio,  # Update session state on slider change
        help="Verplaatst de slider om de verhouding tussen massaminimalisatie en specificatie afwijkingen aan te passen. 0.0 = 1:1, 1.0 = 10:1, 2.0 = 100:1."
    )

    # Calculate penalties based on the ratio
    ratio = 10 ** st.session_state.ratio_log  # Use session state directly
    base_penalty = 10000.0

    # Mass penalty is base_penalty, deviation penalty is base_penalty * ratio
    mass_penalty = base_penalty
    deviation_weight = base_penalty * ratio

    # Ensure penalties stay within reasonable bounds
    mass_penalty = max(0.0, min(mass_penalty, 1000000.0))
    deviation_weight = max(0.0, min(deviation_weight, 1000000.0))

    # Update session state
    st.session_state.mass_penalty = mass_penalty
    st.session_state.deviation_weight = deviation_weight

with st.expander("Veiligheidsmarge", expanded=False):
    # Guard band to stay away from Min/Max bounds
    if 'guard_band_pct' not in st.session_state:
        st.session_state.guard_band_pct = 10.0
    st.session_state.guard_band_pct = st.slider(
        "Veiligheidsmarge t.o.v. Min/Max (%)",
        min_value=0.0,
        max_value=30.0,
        value=float(st.session_state.guard_band_pct),
        step=0.5,
        help="Verschuif de effectieve onder- en bovengrens naar binnen zodat resultaten niet te dicht bij Min/Max komen."
    )

# Reset session state bij wijziging van recept of tank
if ('current_recept' not in st.session_state or st.session_state.current_recept != recept or
    'current_tank' not in st.session_state or st.session_state.current_tank != tank):
    st.session_state.stapgroottes = {grondstof: 0 for grondstof in grondstoffen}
    st.session_state.kosten = {
        grondstof: 0.01 if grondstof == 'Demiwater' else 1.0
        for grondstof in grondstoffen
    }
    st.session_state.excluded_grondstoffen = []
    st.session_state.current_recept = recept
    st.session_state.current_tank = tank

# Stapgroottes sectie
with st.expander("Stapgroottes", expanded=False):
    st.info("Voer de stapgroottes (in kg) in. Stapgrootte 0 betekent geen stappen.")
    for grondstof in grondstoffen:
        step_value = st.number_input(
            f"Stapgrootte (kg) voor {grondstof}",
            value=int(st.session_state.stapgroottes.get(grondstof, 0)),
            step=1,
            min_value=0,
            key=f"step_{recept}_{tank}_{grondstof}",
            help=f"Stapgrootte voor {grondstof} in kg (0 voor geen stappen)."
        )
        st.session_state.stapgroottes[grondstof] = int(step_value)

# Berekeningen zoals in Excel
def bereken_actuele_hoeveelheid_grondstof(row):
    """Berekening kolom G: Actuele hoeveelheid grondstof in kg."""
    if row['Ratio_Element'] == 0:
        return 0.0
    return row['Actuele_Hoeveelheid_Element'] / row['Ratio_Element']

def bereken_toegevoegd_element(row):
    """Berekening kolom I: Toegevoegd element door grondstoftoevoeging."""
    return row['Toevoegen Grondstof'] * row['Ratio_Element']

def bereken_element_concentraties_na_opt(element_groups, toevoegingen_grondstof, totale_massa):
    """Bepaal concentraties per element na optimalisatie."""

    resultaat = {}
    for group in element_groups:
        element_naam = group['element']
        conversie_factor = 100 if str(group['eenheid']).lower() == 'wt%' else 1_000_000
        totale_toevoeging = 0.0
        for ratio, grond_idx in zip(group['ratios'], group['grondstof_indices']):
            coef = float(ratio)
            if coef == 0.0:
                continue
            totale_toevoeging += float(toevoegingen_grondstof[int(grond_idx)]) * coef
        totale_element = float(group['actuele_mass']) + totale_toevoeging
        if totale_massa > 0:
            resultaat[element_naam] = (totale_element / totale_massa) * conversie_factor
        else:
            resultaat[element_naam] = 0.0

    resultaat.setdefault('Geen', 0.0)
    resultaat.setdefault('nan', 0.0)
    return resultaat

# Bereken actuele hoeveelheid element (kolom F)
df['Actuele_Hoeveelheid_Element'] = df['Element'].map(element_actuele_mass_map).fillna(0.0)

# Bereken actuele hoeveelheid grondstof (kolom G)
df['Actuele_Hoeveelheid_Grondstof'] = df.apply(bereken_actuele_hoeveelheid_grondstof, axis=1)

def optimize_toevoegingen(
    massa_product,
    element_groups,
    use_max_volume,
    max_volume,
    grondstoffen,
    stapgroottes_grondstof,
    kosten_grondstof,
    mass_penalty,
    deviation_weight,
    guard_band_pct,
    excluded_grondstoffen=None,
    fixed_grondstoffen=None,
    refine=False,
):
    if excluded_grondstoffen is None:
        excluded_grondstoffen = []
    if fixed_grondstoffen is None:
        fixed_grondstoffen = {}
    if massa_product <= 0:
        return np.zeros(len(grondstoffen), dtype=float), massa_product, "Geen massa; geen optimalisatie nodig."

    excluded_grondstoffen = set(int(idx) for idx in excluded_grondstoffen)
    fixed_grondstoffen = {int(k): float(v) for k, v in fixed_grondstoffen.items()}

    prob = LpProblem("Minimize_Additions_And_Deviations", LpMinimize)

    toevoegingen_grondstof: dict[int, LpVariable | LpAffineExpression | float] = {}
    for g_idx, grondstof in enumerate(grondstoffen):
        if g_idx in excluded_grondstoffen:
            toevoegingen_grondstof[g_idx] = 0.0
            continue

        if g_idx in fixed_grondstoffen:
            toevoegingen_grondstof[g_idx] = fixed_grondstoffen[g_idx]
            continue

        stapgrootte = float(stapgroottes_grondstof[g_idx]) if len(stapgroottes_grondstof) > g_idx else 0.0
        if refine:
            toevoegingen_grondstof[g_idx] = LpVariable(f"add_{g_idx}", lowBound=0, cat="Continuous")
        else:
            if stapgrootte > 0:
                stappen = LpVariable(f"steps_{g_idx}", lowBound=0, cat="Integer")
                toevoegingen_grondstof[g_idx] = stappen * stapgrootte
            else:
                toevoegingen_grondstof[g_idx] = LpVariable(f"add_{g_idx}", lowBound=0, cat="Integer")

    var_toevoegingen = {
        g_idx: expr
        for g_idx, expr in toevoegingen_grondstof.items()
        if isinstance(expr, (LpVariable, LpAffineExpression))
    }
    fixed_add_mass = sum(
        float(expr) for expr in toevoegingen_grondstof.values() if isinstance(expr, (float, int))
    )

    totale_massa_expr = massa_product + fixed_add_mass + lpSum(var_toevoegingen.values())
    effective_massa = massa_product + fixed_add_mass

    diagnostics: list[str] = []
    target_groups: list[int] = []
    targets: dict[int, float] = {}

    def build_element_expression(group: dict[str, object]):
        terms: list = []
        for ratio, grond_idx in zip(group['ratios'], group['grondstof_indices']):
            coef = float(ratio)
            if coef == 0:
                continue
            expr = toevoegingen_grondstof[int(grond_idx)]
            terms.append(coef * expr)
        return lpSum(terms) if terms else 0.0

    for idx, group in enumerate(element_groups):
        conversie_factor = 100 if str(group['eenheid']).lower() == "wt%" else 1_000_000
        fixed_contrib = sum(
            fixed_grondstoffen.get(int(g_idx), 0.0) * float(ratio)
            for ratio, g_idx in zip(group['ratios'], group['grondstof_indices'])
        )
        huidige_mass = float(group['actuele_mass']) + fixed_contrib
        current_conc = (huidige_mass / effective_massa) * conversie_factor if effective_massa > 0 else 0.0

        spec_val = float(group['spec'])
        min_val = float(group['min'])
        max_val = float(group['max'])

        if current_conc < spec_val:
            target_groups.append(idx)
            target_val = (min_val + spec_val) / 2.0
            targets[idx] = target_val
            diag_msg = f"mid tussen Min {min_val:.2f} en Spec {spec_val:.2f}"
        elif current_conc > spec_val:
            target_groups.append(idx)
            target_val = (spec_val + max_val) / 2.0
            targets[idx] = target_val
            diag_msg = f"mid tussen Spec {spec_val:.2f} en Max {max_val:.2f}"
        else:
            target_val = None
            diag_msg = f"Op of gelijk aan Spec, geen target (huidig: {current_conc:.2f})"

        target_str = f"{target_val:.2f}" if target_val is not None else "None"
        diagnostics.append(
            f"{group['element']} target: {target_str} ({diag_msg}, huidig: {current_conc:.2f})"
        )

    total_add_expr = fixed_add_mass + lpSum(var_toevoegingen.values())
    objective = lpSum(
        kosten_grondstof[g_idx] * toevoegingen_grondstof[g_idx] for g_idx in range(len(grondstoffen))
    )
    objective += mass_penalty * total_add_expr

    for idx in target_groups:
        group = element_groups[idx]
        conversie_factor = 100 if str(group['eenheid']).lower() == "wt%" else 1_000_000
        element_expr = build_element_expression(group)
        totale_element = element_expr + float(group['actuele_mass'])
        target_element_mass = targets[idx] * totale_massa_expr / conversie_factor
        dev_plus = LpVariable(f"dev_plus_{idx}", lowBound=0)
        dev_minus = LpVariable(f"dev_minus_{idx}", lowBound=0)
        prob += totale_element == target_element_mass + dev_plus - dev_minus, f"target_dev_{idx}"
        objective += deviation_weight * (dev_plus + dev_minus)

    for idx, group in enumerate(element_groups):
        conversie_factor = 100 if str(group['eenheid']).lower() == "wt%" else 1_000_000
        element_expr = build_element_expression(group)
        totale_element = element_expr + float(group['actuele_mass'])
        band = float(group['max'] - group['min'])
        margin = max(0.0, min(float(guard_band_pct), 49.0)) / 100.0 * band
        min_tight = float(group['min']) + margin
        max_tight = float(group['max']) - margin
        if min_tight > max_tight:
            mid = (float(group['min']) + float(group['max'])) / 2.0
            min_tight = max_tight = mid
        min_element_mass = min_tight * totale_massa_expr / conversie_factor
        max_element_mass = max_tight * totale_massa_expr / conversie_factor
        prob += totale_element >= min_element_mass, f"Min_{idx}"
        prob += totale_element <= max_element_mass, f"Max_{idx}"

    prob += objective

    if use_max_volume and max_volume is not None and max_volume > 0:
        effective_initial_volume = effective_massa / dichtheid
        if effective_initial_volume > max_volume:
            diagnostics.append(
                f"Effectieve initiële massa ({effective_massa:.2f} kg) resulteert in een volume van {effective_initial_volume:.2f} liter, "
                f"wat de maximale volumebeperking van {max_volume:.2f} liter overschrijdt."
            )
        prob += totale_massa_expr / dichtheid <= max_volume, "Max_Volume_Constraint"

    # Limit solver wall-clock time to 25 seconds to avoid long runs
    prob.solve(PULP_CBC_CMD(timeLimit=25, msg=1))

    if LpStatus[prob.status] == "Optimal":
        optimized_toevoegingen = []
        for g_idx in range(len(grondstoffen)):
            expr = toevoegingen_grondstof[g_idx]
            if isinstance(expr, (LpVariable, LpAffineExpression)):
                value = expr.value() if expr.value() is not None else 0.0
                if not refine and stapgroottes_grondstof[g_idx] > 0:
                    value = round(value)
            else:
                value = float(expr)
            optimized_toevoegingen.append(value)

        optimized_toevoegingen = np.array(optimized_toevoegingen, dtype=float)
        optimized_totale_massa = massa_product + float(np.sum(optimized_toevoegingen))

        for diag in diagnostics:
            print(diag)
        print("\nMaterial Additions:")
        for g_idx, value in enumerate(optimized_toevoegingen):
            if value > 0:
                kosten = kosten_grondstof[g_idx] if len(kosten_grondstof) > g_idx else 0.0
                print(f"{grondstoffen[g_idx]}: {value} kg (Cost: {kosten:.2f})")

        return optimized_toevoegingen, optimized_totale_massa, "Succes"

    fail_msg = f"Optimalisatie mislukt: {LpStatus[prob.status]}. Targets:\n" + "\n".join(diagnostics)
    if use_max_volume and max_volume is not None and max_volume > 0:
        initial_volume = massa_product / dichtheid
        fail_msg += (
            f"\nControleer de maximale volumebeperking: huidige massa ({massa_product:.2f} kg) heeft een volume van "
            f"{initial_volume:.2f} liter, terwijl het maximum {max_volume:.2f} liter is."
        )
    return None, None, fail_msg
    
# Initialisatie van session state
if 'excluded_grondstoffen' not in st.session_state:
    st.session_state.excluded_grondstoffen = []
if 'optimized_resultaten' not in st.session_state:
    st.session_state.optimized_resultaten = None
if 'optimized_status' not in st.session_state:
    st.session_state.optimized_status = None
if 'optimized_totale_massa' not in st.session_state:
    st.session_state.optimized_totale_massa = None
if 'optimized_toevoegingen_grond' not in st.session_state:
    st.session_state.optimized_toevoegingen_grond = None
if 'optimized_toevoegingen_rows' not in st.session_state:
    st.session_state.optimized_toevoegingen_rows = None

# Knop voor optimalisatie
st.header("Geoptimaliseerde Toevoegingen")
if st.button("Bereken Geoptimaliseerde Toevoegingen", help="Klik om de optimalisatie te starten op basis van de ingevoerde concentraties."):
    with st.spinner("Optimalisatie wordt uitgevoerd..."):
        # Clear previous optimization results to avoid stale data
        st.session_state.optimized_resultaten = None
        st.session_state.optimized_status = None
        st.session_state.optimized_totale_massa = None
        st.session_state.optimized_toevoegingen_grond = None
        st.session_state.optimized_toevoegingen_rows = None

        stapgroottes_grondstof = np.array(
            [st.session_state.stapgroottes.get(grondstof, 0) for grondstof in grondstoffen],
            dtype=float,
        )
        kosten_grondstof = np.array(
            [st.session_state.kosten.get(grondstof, 1.0) for grondstof in grondstoffen],
            dtype=float,
        )

        optimized_toevoegingen_grond, optimized_totale_massa, status = optimize_toevoegingen(
            massa_product,
            element_groups,
            use_max_volume,
            max_volume_liters,
            grondstoffen,
            stapgroottes_grondstof,
            kosten_grondstof,
            st.session_state.mass_penalty,
            st.session_state.deviation_weight,
            st.session_state.guard_band_pct,
            excluded_grondstoffen=[],  # Geen exclusies voor originele optimalisatie
        )
        if status == "Succes":
            # Check for small additions to refine
            refine_indices = [i for i, value in enumerate(optimized_toevoegingen_grond) if 0 < value <= 9]
            if refine_indices:
                fixed_values = {
                    i: optimized_toevoegingen_grond[i]
                    for i in range(len(optimized_toevoegingen_grond))
                    if i not in refine_indices
                }
                optimized_toevoegingen_grond, optimized_totale_massa, status = optimize_toevoegingen(
                    massa_product,
                    element_groups,
                    use_max_volume,
                    max_volume_liters,
                    grondstoffen,
                    stapgroottes_grondstof,
                    kosten_grondstof,
                    st.session_state.mass_penalty,
                    st.session_state.deviation_weight,
                    st.session_state.guard_band_pct,
                    excluded_grondstoffen=[],
                    fixed_grondstoffen=fixed_values,
                    refine=True,
                )
            if status == "Succes":
                toevoegingen_per_rij = optimized_toevoegingen_grond[row_to_grondstof_idx]
                df['Toevoegen Grondstof'] = toevoegingen_per_rij
                df['Geoptimaliseerde_Toegevoegd_Element'] = df.apply(bereken_toegevoegd_element, axis=1)
                na_opt_map = bereken_element_concentraties_na_opt(
                    element_groups,
                    optimized_toevoegingen_grond,
                    optimized_totale_massa,
                )
                df['Na Optimalisatie'] = df['Element'].map(lambda e: na_opt_map.get(e, 0.0))
                df['Binnen Specs'] = (df['Na Optimalisatie'] >= df['Min']) & (df['Na Optimalisatie'] <= df['Max'])
                st.session_state.optimized_resultaten = df[['Grondstof', 'Element', 'Gemeten Concentratie', 'Toevoegen Grondstof', 'Na Optimalisatie', 'Spec', 'Min', 'Max', 'Binnen Specs', 'Eenheid']].copy()
                st.session_state.optimized_status = status
                st.session_state.optimized_totale_massa = optimized_totale_massa
                st.session_state.optimized_toevoegingen_grond = optimized_toevoegingen_grond
                st.session_state.optimized_toevoegingen_rows = toevoegingen_per_rij
        if status != "Succes":
            st.error(f"Optimalisatie mislukt: {status}")
            if "Infeasible" in status and use_max_volume:
                st.warning(
                    f"De maximale volumebeperking ({max_volume_liters:.2f} liter) is waarschijnlijk te streng. "
                    "Probeer het maximale volume te verhogen of schakel de volumebeperking uit."
                )
            else:
                st.warning(
                    "De optimalisatie kon geen oplossing vinden. Controleer de invoerparameters (bijv. gemeten concentraties, specificaties, of stapgroottes) en probeer opnieuw."
                )

# Toon originele resultaten als ze beschikbaar zijn
if st.session_state.optimized_status == "Succes" and st.session_state.optimized_resultaten is not None:
    def highlight_specs(val):
        color = 'green' if val else 'red'
        return f'background-color: {color}'
    
    def format_number(val, eenheid):
        if eenheid == 'mg/kg' or val == 0:
            return '{:.0f}'.format(val)
        return '{:.3f}'.format(val).rstrip('0').rstrip('.')
    
    def format_toevoegen_grondstof(val):
        try:
            if val is None:
                return ''
            if isinstance(val, float) and np.isnan(val):
                return ''
            v = float(val)
        except (TypeError, ValueError):
            return ''
        if v < 10:
            return '{:.1f}'.format(v)
        return '{:.0f}'.format(v)
    
    format_dict = {
        'Toevoegen Grondstof': format_toevoegen_grondstof,
        'Gemeten Concentratie': lambda x: '{:.2f}'.format(x),
        'Na Optimalisatie': lambda x: '{:.2f}'.format(x),
        'Spec': lambda x: format_number(x, st.session_state.optimized_resultaten.loc[st.session_state.optimized_resultaten['Spec'] == x, 'Eenheid'].iloc[0] if not st.session_state.optimized_resultaten.loc[st.session_state.optimized_resultaten['Spec'] == x, 'Eenheid'].empty else 'wt%'),
        'Min': lambda x: format_number(x, st.session_state.optimized_resultaten.loc[st.session_state.optimized_resultaten['Min'] == x, 'Eenheid'].iloc[0] if not st.session_state.optimized_resultaten.loc[st.session_state.optimized_resultaten['Min'] == x, 'Eenheid'].empty else 'wt%'),
        'Max': lambda x: format_number(x, st.session_state.optimized_resultaten.loc[st.session_state.optimized_resultaten['Max'] == x, 'Eenheid'].iloc[0] if not st.session_state.optimized_resultaten.loc[st.session_state.optimized_resultaten['Max'] == x, 'Eenheid'].empty else 'wt%')
    }
    
    totale_toevoeging = st.session_state.optimized_totale_massa - massa_product
    percentage_toename = (totale_toevoeging / massa_product) * 100 if massa_product > 0 else 0
    st.success(f"Optimalisatie succesvol! Totale massa na correctie: **{st.session_state.optimized_totale_massa:.2f} kg** (toevoeging: **{totale_toevoeging:.2f} kg**, {percentage_toename:.2f}% toename)")
    
    def show_grouped_table(df_src):
        res = df_src.copy()

        # Count sources per element and detect duplicates
        elem_source_count = res.groupby('Element')['Grondstof'].transform('nunique')
        elem_is_duplicate = elem_source_count > 1

        # Does a source contribute to more than one element?
        grondstof_multi_map = res.groupby('Grondstof')['Element'].nunique().gt(1).to_dict()
        src_is_multi = res['Grondstof'].map(grondstof_multi_map).fillna(False)

        # For each element: are ALL its contributing sources multi-element?
        all_multi_by_elem = (
            res.assign(_is_multi=src_is_multi)
               .groupby('Element')['_is_multi']
               .transform('all')
        )

        # Elements to combine visually: duplicates AND all sources are multi-element
        combine_flag = elem_is_duplicate & all_multi_by_elem

        # Keep only first row for combined elements; keep all rows otherwise
        keep_mask = (~combine_flag) | (~res.duplicated(subset=['Element'], keep='first'))
        res = res.loc[keep_mask].copy()

        # After filtering, blank Grondstof and Toevoegen for the remaining rows of combined elements
        combine_rows_after = combine_flag.loc[res.index]
        res.loc[combine_rows_after, 'Grondstof'] = ''
        res.loc[combine_rows_after, 'Toevoegen Grondstof'] = np.nan

        st.dataframe(
            res.style
              .applymap(highlight_specs, subset=['Binnen Specs'])
              .format(format_dict, na_rep='')
        )

    # original results
    show_grouped_table(st.session_state.optimized_resultaten)
    
    st.subheader("Waarschuwingen")
    warnings_present = False
    for index, row in st.session_state.optimized_resultaten.iterrows():
        if not row['Binnen Specs'] and row['Element'] != 'Geen':
            st.warning(f"{row['Element']} ({row['Na Optimalisatie']:.2f} {row['Eenheid']}) ligt buiten specificaties ({row['Min']} - {row['Max']})")
            warnings_present = True
    if not warnings_present:
        st.success("Alle concentraties liggen binnen de specificaties!")
    
    # Export button for original results (Excel only)
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        st.session_state.optimized_resultaten.to_excel(writer, index=False, sheet_name='Resultaten')
    excel_data = output.getvalue()
    st.download_button(
        label="Download Resultaten als Excel",
        data=excel_data,
        file_name=f'geoptimaliseerde_correctie_{recept}_{tank}.xlsx',
        mime='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        help="Download de geoptimaliseerde resultaten als Excel-bestand."
    )

    # Samenvatting per Grondstof (uitgeschakeld, behouden voor later gebruik)
    # grondstof_overzicht = []
    # grondstoffen_volgorde = st.session_state.get('current_grondstoffen', grondstoffen)
    # if st.session_state.optimized_toevoegingen_grond is not None:
    #     for g_idx, grondstof in enumerate(grondstoffen_volgorde):
    #         toevoeging = st.session_state.optimized_toevoegingen_grond[g_idx]
    #         matching_rows = st.session_state.optimized_resultaten[
    #             st.session_state.optimized_resultaten['Grondstof'] == grondstof
    #         ]
    #         if matching_rows.empty:
    #             continue
    #         elementen_info = "; ".join(
    #             f"{r['Element']}: {r['Na Optimalisatie']:.2f} {r['Eenheid']}"
    #             for _, r in matching_rows.iterrows()
    #         )
    #         grondstof_overzicht.append(
    #             {
    #                 "Grondstof": grondstof,
    #                 "Toevoegen (kg)": round(toevoeging, 3),
    #                 "Elementen na optimalisatie": elementen_info,
    #             }
    #         )
    # if grondstof_overzicht:
    #     st.subheader("Samenvatting per Grondstof")
    #     st.table(pd.DataFrame(grondstof_overzicht))

# Aangepaste Correctie Sectie
if st.session_state.optimized_toevoegingen_grond is not None:
    excluded_grondstoffen = []
    st.info("Selecteer de grondstoffen die je wilt toevoegen. Vink uit om ze uit te sluiten bij heroptimalisatie.")
    for g_idx, grondstof in enumerate(grondstoffen):
        default_value = st.session_state.get(f"exclude_{recept}_{tank}_{grondstof}", True)
        label_suffix = ""
        huidige_toevoeging = st.session_state.optimized_toevoegingen_grond[g_idx]
        if huidige_toevoeging > 0:
            label_suffix = f" ({huidige_toevoeging:.2f} kg)"
        toevoegen = st.checkbox(
            f"{grondstof} toevoegen{label_suffix}",
            value=default_value,
            key=f"exclude_{recept}_{tank}_{grondstof}"
        )
        if not toevoegen:
            excluded_grondstoffen.append(g_idx)

    st.session_state.excluded_grondstoffen = excluded_grondstoffen

    if st.button(
        "Herbereken Zonder Geselecteerde Grondstoffen",
        help="Herbereken optimalisatie zonder de uitgeschakelde grondstoffen",
    ):
        with st.spinner("Aangepaste optimalisatie wordt uitgevoerd..."):
            adj_df = df.copy()
            stapgroottes_grondstof = np.array(
                [st.session_state.stapgroottes.get(grondstof, 0) for grondstof in grondstoffen],
                dtype=float,
            )
            kosten_grondstof = np.array(
                [st.session_state.kosten.get(grondstof, 1.0) for grondstof in grondstoffen],
                dtype=float,
            )

            adj_optimized_toevoegingen, adj_optimized_totale_massa, adj_status = optimize_toevoegingen(
                massa_product,
                element_groups,
                use_max_volume,
                max_volume_liters,
                grondstoffen,
                stapgroottes_grondstof,
                kosten_grondstof,
                st.session_state.mass_penalty,
                st.session_state.deviation_weight,
                st.session_state.guard_band_pct,
                excluded_grondstoffen=st.session_state.excluded_grondstoffen,
            )
            if adj_status == "Succes":
                # Check for small additions to refine
                refine_indices = [
                    i for i, value in enumerate(adj_optimized_toevoegingen) if 0 < value <= 9
                ]
                if refine_indices:
                    fixed_values = {
                        i: adj_optimized_toevoegingen[i]
                        for i in range(len(adj_optimized_toevoegingen))
                        if i not in refine_indices
                    }
                    adj_optimized_toevoegingen, adj_optimized_totale_massa, adj_status = optimize_toevoegingen(
                        massa_product,
                        element_groups,
                        use_max_volume,
                        max_volume_liters,
                        grondstoffen,
                        stapgroottes_grondstof,
                        kosten_grondstof,
                        st.session_state.mass_penalty,
                        st.session_state.deviation_weight,
                        st.session_state.guard_band_pct,
                        excluded_grondstoffen=st.session_state.excluded_grondstoffen,
                        fixed_grondstoffen=fixed_values,
                        refine=True,
                    )
            if adj_status == "Succes":
                adj_toevoegingen_per_rij = adj_optimized_toevoegingen[row_to_grondstof_idx]
                adj_df['Toevoegen Grondstof'] = adj_toevoegingen_per_rij
                adj_df['Geoptimaliseerde_Toegevoegd_Element'] = adj_df.apply(bereken_toegevoegd_element, axis=1)
                adj_na_opt_map = bereken_element_concentraties_na_opt(
                    element_groups,
                    adj_optimized_toevoegingen,
                    adj_optimized_totale_massa,
                )
                adj_df['Na Optimalisatie'] = adj_df['Element'].map(lambda e: adj_na_opt_map.get(e, 0.0))
                adj_df['Binnen Specs'] = (adj_df['Na Optimalisatie'] >= adj_df['Min']) & (adj_df['Na Optimalisatie'] <= adj_df['Max'])
                st.session_state.optimized_toevoegingen_grond = adj_optimized_toevoegingen
                st.session_state.optimized_toevoegingen_rows = adj_toevoegingen_per_rij
                adj_totale_toevoeging = adj_optimized_totale_massa - massa_product
                adj_percentage_toename = (adj_totale_toevoeging / massa_product) * 100 if massa_product > 0 else 0
                st.success(f"Aangepaste optimalisatie succesvol! Totale massa na correctie: **{adj_optimized_totale_massa:.2f} kg** (toevoeging: **{adj_totale_toevoeging:.2f} kg**, {adj_percentage_toename:.2f}% toename)")
                
                adj_optimized_resultaten = adj_df[['Grondstof', 'Element', 'Gemeten Concentratie', 'Toevoegen Grondstof', 'Na Optimalisatie', 'Spec', 'Min', 'Max', 'Binnen Specs', 'Eenheid']]
                # adjusted results
                if 'show_grouped_table' not in globals():
                    def show_grouped_table(df_src):
                        res = df_src.copy()

                        elem_source_count = res.groupby('Element')['Grondstof'].transform('nunique')
                        elem_is_duplicate = elem_source_count > 1

                        grondstof_multi_map = res.groupby('Grondstof')['Element'].nunique().gt(1).to_dict()
                        src_is_multi = res['Grondstof'].map(grondstof_multi_map).fillna(False)

                        all_multi_by_elem = (
                            res.assign(_is_multi=src_is_multi)
                               .groupby('Element')['_is_multi']
                               .transform('all')
                        )

                        combine_flag = elem_is_duplicate & all_multi_by_elem

                        keep_mask = (~combine_flag) | (~res.duplicated(subset=['Element'], keep='first'))
                        res = res.loc[keep_mask].copy()

                        combine_rows_after = combine_flag.loc[res.index]
                        res.loc[combine_rows_after, 'Grondstof'] = ''
                        res.loc[combine_rows_after, 'Toevoegen Grondstof'] = np.nan

                        st.dataframe(
                            res.style
                              .applymap(highlight_specs, subset=['Binnen Specs'])
                              .format(format_dict, na_rep='')
                        )

                show_grouped_table(adj_optimized_resultaten)
                
                st.subheader("Aangepaste Waarschuwingen")
                warnings_present = False
                for index, row in adj_optimized_resultaten.iterrows():
                    if not row['Binnen Specs'] and row['Element'] != 'Geen':
                        st.warning(f"{row['Element']} ({row['Na Optimalisatie']:.2f} {row['Eenheid']}) ligt buiten specificaties ({row['Min']} - {row['Max']})")
                        warnings_present = True
                if not warnings_present:
                    st.success("Alle concentraties liggen binnen de specificaties!")
                
                # Export button for adjusted results (Excel only)
                output = io.BytesIO()
                with pd.ExcelWriter(output, engine='openpyxl') as writer:
                    adj_optimized_resultaten.to_excel(writer, index=False, sheet_name='Aangepaste_Resultaten')
                excel_data = output.getvalue()
                st.download_button(
                    label="Download Aangepaste Resultaten als Excel",
                    data=excel_data,
                    file_name=f'aangepaste_correctie_{recept}_{tank}.xlsx',
                    mime='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                    help="Download de aangepaste geoptimaliseerde resultaten als Excel-bestand."
                )
            else:
                st.error(f"Aangepaste optimalisatie mislukt: {adj_status}")
else:
    st.info("Geen optimalisatieresultaten beschikbaar. Voer eerst de originele optimalisatie uit.")
